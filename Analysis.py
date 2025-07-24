# Author: Zedaine McDonald

from openpyxl import load_workbook
from io import BytesIO
import pandas as pd

def parse_budget(file):
    wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
    sheet = wb.active
    month_headers = [cell.value for cell in sheet[1][1:14]]  # B1 to N1

    data = []
    current_category = None

    for row in sheet.iter_rows(min_row=2):
        name_cell = row[0]
        name = name_cell.value
        if name is None:
            continue

        fill = name_cell.fill
        is_category = (
            fill and fill.start_color and fill.start_color.type == "rgb" and
            fill.start_color.rgb in ["FFFFFF00", "FFFF00", "FFFFFF99"]
        )

        if is_category:
            current_category = name
            continue

        months = [cell.value for cell in row[1:13]]  # Month 1–12
        total = row[13].value  # Total column

        row_data = {
            "Category": current_category,
            "Subcategory": name,
            **{month_headers[i]: months[i] for i in range(len(months))},
            "Total": total
        }
        data.append(row_data)

    return pd.DataFrame(data)


def parse_expense(file):
    df = pd.read_excel(file)
    df["Date"] = pd.to_datetime(df["Date"])
    df["Month"] = df["Date"].dt.strftime("%B")
    return df


def category_summary(expense_df, budget_df):
    cat_total = expense_df.groupby("Category").agg(
        Total_Spent=("Amount", "sum"),
        Purchase_Count=("Amount", "count")
    ).reset_index()

    budget_total = budget_df.groupby("Category").agg(
        Budgeted_Amount=("Total", "sum")
    ).reset_index()

    merged = pd.merge(cat_total, budget_total, on="Category", how="left")
    merged["Variance"] = merged["Budgeted_Amount"] - merged["Total_Spent"]
    return merged


def subcategory_summary(expense_df, budget_df):
    sub_total = expense_df.groupby(["Category", "Sub-Category"]).agg(
        Total_Spent=("Amount", "sum")
    ).reset_index()

    budget_sub = budget_df.groupby(["Category", "Subcategory"]).agg(
        Budgeted_Amount=("Total", "sum")
    ).reset_index()

    merged = pd.merge(sub_total, budget_sub, left_on=["Category", "Sub-Category"],
                      right_on=["Category", "Subcategory"], how="left")
    merged["Variance"] = merged["Budgeted_Amount"] - merged["Total_Spent"]
    return merged


def monthly_summary(expense_df):
    return expense_df.groupby("Month").agg(
        Total_Spent=("Amount", "sum")
    ).reset_index()


def vendor_summary(expense_df):
    return expense_df.groupby("Vendor").agg(
        Total_Spent=("Amount", "sum"),
        Purchase_Count=("Amount", "count")
    ).reset_index()
